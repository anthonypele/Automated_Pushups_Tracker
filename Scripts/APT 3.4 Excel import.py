# %% Load libraries
import tkinter as tk
from tkinter import filedialog
import re
import datetime
import os
import openpyxl.styles
import pandas as pd
import openpyxl
from openpyxl.styles import numbers 
import calendar
import pyperclip
import numpy as np  

# %% Open a file dialog to choose a file
def choose_file():
    # Create a hidden root window (we don't need it to show up)
    root = tk.Tk()
    root.withdraw() # Hides the Tkinter root window

    # Open the file dialog and return the selected file path
    file_path = filedialog.askopenfilename(title='Select the WA data file', filetypes=[('Textfiles', '*.txt')], initialdir=r"C:\Users\Anthony\YandexDisk\_Programming\APT\Data")

    # Return the selected file path
    return file_path

# Read my txt file
with open(choose_file(), 'r', encoding="utf-8") as file:
    lines = file.readlines() #creating a list made of each line from the txt imported WA file

pattern = r"\[(\d{2}\.\d{2}\.\d{4}) (\d{1,2}:\d{2})\] (.*?): (\d+)"

aliases = {
    "Michael Ice&Fire Perm": "Миша",
    "Anthony": "Антон",
    "Гриша Соловьев": "Гриша",
    "Павел Антонюк РТ": "Паша",
    "Андрей Палыч Павлов": "Палыч",
    "Роман Аландаров": "Рома",
    "Андрей Русанов": "Андрей"
}

# New variables
new_lines = []
data = []

# Changing names
def replace_name(match):
    date, time, name, pushups = match.groups()
    new_name = aliases.get(name, name)
    return f'[{date} {time}] {new_name}: {pushups}'

# Changing my lines
new_lines = [re.sub(pattern, replace_name, line) for line in lines]

# Parse lines
for line in new_lines:
    match = re.match(pattern, line.strip())
    if match:
        date, time, name, pushups = match.groups()
        data.append({
            'date': date,
            'time': time,
            'name': name,
            'pushups': int(pushups)
        })

# Create a DataFrame 
df = pd.DataFrame(data)

# ======================
# Creating pivot table
# ======================

# %% Calculate the last 2 months - using other way around below
# today = datetime.date.today()
# first_day_this_month = today.replace(day=1) # Gets the first day of this month
# last_month_end = first_day_this_month - datetime.timedelta(days=1) # Goes back 1 day
# second_last_month_end = (first_day_this_month - datetime.timedelta(days=1)).replace(day=1) - datetime.timedelta(days=1)

# last_month = calendar.month_abbr[last_month_end.month].lower() ##!! It seems that I don't have calendar library installed!!
# second_last_month = calendar.month_abbr[second_last_month_end.month].lower()

# print(last_month)
# print(second_last_month)

# Ensure the date column is treated as data
df['date'] = pd.to_datetime(df['date'], format="%d.%m.%Y")

# Create a new column for the month
df['month'] = df['date'].dt.to_period('M')

# Create the pivot table
pivot = pd.pivot_table(
    df,
    index='name',
    columns='month',
    values='pushups',
    aggfunc='sum',
    fill_value=0,
    margins=True,
    margins_name='Итог'
)

# Remove the total column, that calculates rows
pivot = pivot.drop(columns='Итог')

# Get the list of all month columns in the pivot
all_months = [col for col in pivot.columns] #if col != 'Итог']
# Sort them
all_months_sorted = sorted(all_months)
# Choose the last 2 months only
last_two_months = all_months_sorted[-2:]
# Filter the pivot for only these 2 months
filtered_pivot = pivot[last_two_months].copy()

# Create % column to see the difference in %
filtered_pivot['%'] = (
    (filtered_pivot[last_two_months[-1]] / filtered_pivot[last_two_months[-2]]) - 1 # division on 0 is treated in pivot2.
).round(2) 
# Sort the pivot table the way I want
if 'Итог' in filtered_pivot.index:
    total_row = filtered_pivot.loc[['Итог']] # Keep as DataFrame
    filtered_pivot = filtered_pivot.drop('Итог')
else:
    total_row = None
filtered_pivot = filtered_pivot.sort_values([last_two_months[-1]], ascending=False) 
if total_row is not None:
    filtered_pivot = pd.concat([filtered_pivot, total_row])

#pyperclip.copy(filtered_pivot)

# Creating other pivot table with counts to test it
# Create the pivot table
pivot2 = pd.pivot_table(
    df,
    index='name',
    columns='month',
    values='pushups',
    aggfunc='count',
    fill_value=0,
    margins=True,
    margins_name='Итог'
)

# Remove the total column, that calculates rows
pivot2 = pivot2.drop(columns='Итог')
# Filter the pivot for only these 2 months
filtered_pivot2 = pivot2[last_two_months].copy()

# Create % column to see the difference in %
division = filtered_pivot2[last_two_months[-1]] / filtered_pivot2[last_two_months[-2]]
filtered_pivot2['%'] = np.where( # A dance to get rid of division on 0 
    np.isfinite(division), # Condition
    (division - 1).round(2), # If True
    np.nan                      # If False
)
# Sort the pivot table the way I want
if 'Итог' in filtered_pivot2.index:
    total_row = filtered_pivot2.loc[['Итог']] # Keep as DataFrame
    filtered_pivot2 = filtered_pivot2.drop('Итог')
else:
    total_row = None
filtered_pivot2 = filtered_pivot2.sort_values([last_two_months[-1]], ascending=False)
if total_row is not None:
    filtered_pivot2 = pd.concat([filtered_pivot2, total_row])


# %% ==================
# Saving the file
# ==================
# Get current timestamp 
now = datetime.datetime.now()
timestamp = now.strftime('%Y-%m-%d_%H-%M-%S') # Format: 2025-04-10_15-42-07

# Create filename with timestamp
filename = f"pushups_data_{timestamp}.xlsx"

# Define save folder
save_folder = r"C:\Users\Anthony\YandexDisk\_Programming\APT\Data"

# Full path to save the file
file_path = os.path.join(save_folder, filename)

# Save to excel - This is to save only 1 sheet. For multilple use panda
#pivot.to_excel(file_path)

# Save multiple sheets to excel
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Data', index=False)
    filtered_pivot2.to_excel(writer, sheet_name='Count')
    filtered_pivot.to_excel(writer, sheet_name='Last_2_months')

# Style the data
    workbook = writer.book
    sheet = writer.sheets['Last_2_months']
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=sheet.max_column)
        cell.number_format = '0%'
    sheet = writer.sheets['Count']
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=sheet.max_column)
        cell.number_format = '0%'

# 
print("✅ Data successfully exported to {file path}")

os.startfile(file_path)

